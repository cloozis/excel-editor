import openpyxl
import re

# Чтение исходного файла
wb = openpyxl.load_workbook('export_base_structure.xlsx')
sheet = wb.active

# Теги которые ищем и заменяем
base_slug = {
    'Мапикс':'Тест',
    'Ртс телеком':'Тест',
    'оборудование':'Тест',
    'Ultra Distribution Company':'Тест',
    'Я на связи':'Тест',
    'Айс Партнерс':'Тест',
    'телеком':'Тест',
    'Distribution':'Тест',
    'None':'Тест',
    'info@export-base.ru':'Тест',
    'техподдержка, тел. 8 800 775-29-12':'Тест',
}

# перебор строки по тегам
def subUpdate(data):

    data = str(data)
    count = len(base_slug)    

    if count > 0:
        for key in base_slug.keys():
            key = str(key)
            base_slug_str = str(base_slug[key])
            data = re.sub(key, base_slug_str, data)
            # data = data.replace(key, base_slug_str)     

    return data
    

# запуск переборки строки
def check(data):
    return subUpdate(data)


# Проверка и изменение данных в первой колонке
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
    cell_value = row[0].value
    per = check(cell_value)
    print(per)
    # Проверка условия
    if per:
        # Изменение значения
        row[0].value = per

# Сохранение изменений в новый файл
wb.save('output.xlsx')